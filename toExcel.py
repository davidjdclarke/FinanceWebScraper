from openpyxl import load_workbook

def toXl(filename, pos, x):

    wb = load_workbook(filename=filename)
    ws = wb.active

    ws[pos] = x

    wb.save(filename)

    print(str(ws[pos].value) + " saved to position " + pos)
    wb.close()


if __name__ == "__main__":
    url = 'C:\\Users\\David Clarke\\Desktop\\Book1.xlsx'
    toXl(url, 'A10', 55)
